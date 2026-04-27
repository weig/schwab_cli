"""Index member fetchers.

stockanalysis.com is the primary provider — its public list pages
(``/list/sp-500-stocks/`` etc.) render one ``<a href="/stocks/SYM/">``
per member, which is stable enough to scrape. SSGA is the fallback
for SPX and DJI only (their public xlsx holdings of SPY / DIA).

Both providers return ``set[str]`` of normalized symbols (Schwab's
dot form, e.g. ``'BRK.B'``).
"""
from __future__ import annotations

import io
import re

import httpx


INDEX_TO_STOCKANALYSIS_SLUG = {
    "SPX": "sp-500",
    "DJI": "dow-jones",
    "NQ":  "nasdaq-100",
}

INDEX_TO_SSGA_ETF = {
    "SPX": "spy",
    "DJI": "dia",
}

_STOCKANALYSIS_BASE = "https://stockanalysis.com/list"
_SSGA_URL = (
    "https://www.ssga.com/us/en/intermediary/library-content/products/"
    "fund-data/etfs/us/holdings-daily-us-en-{etf}.xlsx"
)
# stockanalysis.com / Cloudflare reject our identifier-style UA. A
# vanilla browser UA goes through cleanly and is what every other
# scraper uses anyway.
_BROWSER_UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/126.0.0.0 Safari/537.36"
)


def fetch_stockanalysis_members(
    index_name: str,
    *,
    client: httpx.Client,
) -> set[str]:
    """Pull members from stockanalysis.com CSV download."""
    slug = INDEX_TO_STOCKANALYSIS_SLUG.get(index_name)
    if slug is None:
        raise ValueError(
            f"{index_name!r} not supported by stockanalysis.com "
            f"(supported: {sorted(INDEX_TO_STOCKANALYSIS_SLUG)})"
        )
    url = f"{_STOCKANALYSIS_BASE}/{slug}-stocks/"
    resp = client.get(
        url,
        headers={
            "User-Agent": _BROWSER_UA,
            "Accept": "text/html,application/xhtml+xml,*/*",
        },
        follow_redirects=True,
    )
    if resp.status_code != 200:
        raise RuntimeError(
            f"stockanalysis.com HTTP {resp.status_code} for {url}"
        )
    return _parse_stockanalysis_html(resp.text)


# Tickers appear as ``href="/stocks/<lowercase>/"`` links in the
# rendered table. Skipping path words that share the same prefix
# (e.g. ``/stocks/screener/``).
_TICKER_HREF_RE = re.compile(r'href="/stocks/([a-z][a-z0-9.\-]{0,5})/"')
_PATH_WORDS = frozenset({"screener", "compare", "industry", "sector", "list"})


def _parse_stockanalysis_html(html: str) -> set[str]:
    """Extract member tickers from a rendered list page.

    The page renders one ``<a href="/stocks/SYM/">`` per member; we
    collect those, drop the few non-ticker path words, and normalize
    to Schwab's dot form (``brk-b`` → ``BRK.B``).
    """
    out: set[str] = set()
    for raw in _TICKER_HREF_RE.findall(html):
        if raw in _PATH_WORDS:
            continue
        out.add(_normalize_symbol(raw))
    if not out:
        raise RuntimeError(
            "stockanalysis page had no /stocks/<symbol>/ links — "
            "page layout may have changed"
        )
    return out


def _normalize_symbol(s: str) -> str:
    """Schwab uses dots (``BRK.B``); SSGA uses dashes (``BRK-B``).

    stockanalysis.com uses dots already, so this is a no-op for that
    path. Used by both adapters so callers always get the same form.
    """
    return s.replace("-", ".").upper()


def fetch_ssga_members(
    index_name: str,
    *,
    client: httpx.Client,
) -> set[str]:
    """Pull members from SSGA's daily-holdings xlsx (SPY / DIA only)."""
    etf = INDEX_TO_SSGA_ETF.get(index_name)
    if etf is None:
        raise ValueError(
            f"{index_name!r} not supported by SSGA "
            f"(supported: {sorted(INDEX_TO_SSGA_ETF)})"
        )
    url = _SSGA_URL.format(etf=etf)
    # SSGA recently moved their canonical xlsx URL and now serves a
    # 301 from the legacy ``/us/en/intermediary/`` prefix. Always
    # follow redirects so the adapter survives the next move too.
    resp = client.get(
        url,
        headers={"User-Agent": _BROWSER_UA},
        follow_redirects=True,
    )
    if resp.status_code != 200:
        raise RuntimeError(f"SSGA HTTP {resp.status_code} for {url}")
    return _parse_ssga_xlsx(resp.content)


def _parse_ssga_xlsx(blob: bytes) -> set[str]:
    """Walk the SSGA xlsx and collect tickers from the 'Ticker' column."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    ticker_col = None
    for i, row in enumerate(rows):
        if row is None:
            continue
        normalized = [str(c or "").strip().lower() for c in row]
        if "ticker" in normalized:
            header_idx = i
            ticker_col = normalized.index("ticker")
            break
    if header_idx is None or ticker_col is None:
        raise RuntimeError("SSGA xlsx: could not locate Ticker column")
    out: set[str] = set()
    for row in rows[header_idx + 1:]:
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            break  # blank row terminates the table
        sym = row[ticker_col]
        if sym is None:
            continue
        sym = str(sym).strip()
        if not sym or sym.upper() in {"USD", "CASH", "-"}:
            continue
        out.add(_normalize_symbol(sym))
    if not out:
        raise RuntimeError("SSGA xlsx had no parseable tickers")
    return out


def fetch_index_members(
    index_name: str,
    *,
    client: httpx.Client,
) -> set[str]:
    """Resolve members via primary (stockanalysis), then fallback (SSGA)."""
    if index_name == "RUT":
        raise NotImplementedError(
            "RUT (Russell 2000): no clean upstream provider yet — "
            "see docs/plan/dataset.md (TODO)."
        )

    errors: list[str] = []

    if index_name in INDEX_TO_STOCKANALYSIS_SLUG:
        try:
            return fetch_stockanalysis_members(index_name, client=client)
        except (RuntimeError, ValueError) as e:
            errors.append(f"stockanalysis: {e}")

    if index_name in INDEX_TO_SSGA_ETF:
        try:
            return fetch_ssga_members(index_name, client=client)
        except (RuntimeError, ValueError) as e:
            errors.append(f"ssga: {e}")

    raise RuntimeError(
        f"{index_name}: all providers failed — {'; '.join(errors)}"
    )
